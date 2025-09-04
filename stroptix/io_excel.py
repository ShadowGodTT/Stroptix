from __future__ import annotations

from pathlib import Path
from typing import List, Optional

import pandas as pd

from .models import InputModel, MemberCandidate
from .parsing import parse_bay_pattern


INPUT_SHEET_NAME = "Input"
OUTPUT_SHEET_NAME = "StrOptix Output"
MEMBER_TABLE_SHEET = "Member Table"
TEMPLATE_SHEET_NAME = "Output (Template)"


# Base expected fields (we will also map synonyms)
INPUT_COLUMNS = [
    "design_code",
    "frame_type",
    "span_m",
    "eave_height_m",
    "bay_spacing_m",
    "dead_load_kPa",
    "live_load_kPa",
    "collateral_kPa",
    "wind_speed_mps",
    "seismic_zone",
    "steel_grade",
    "Members_count",
]

# Synonym mapping from user-friendly labels to our model fields
SYNONYMS = {
    # Design/code
    "Desing code": "design_code",
    "Design code": "design_code",
    # Frame/building type
    "Building type": "frame_type",
    # Span/length/width
    "Length in m (o/o of steel)": "span_m",
    "Width in m (o/o of steel)": "Width in m (o/o of steel)",
    # Heights
    "Left eave height in m": "eave_height_m",
    "Right eave height in m": "Right eave height in m",
    # Bay spacing
    "Right end wall bay spacing in m": "Right end wall bay spacing in m",
    # Do not remap "Side wall bay spacing"; it directly maps to InputModel via alias
    # Loads
    "Dead Load in kN/m2": "dead_load_kPa",
    "Live Load in kN/m2": "live_load_kPa",
    "Collateral load kN/m2": "collateral_kPa",
    # Wind
    "Max wind speed in m/s": "wind_speed_mps",
    # Seismic
    "Seismic zone  co-efficient (III)": "seismic_zone",
    # Steel grade
    "Steel material grade (Mpa)": "steel_grade",
    # Starter pack snake_case / alternative names
    "width_m": "span_m",
    "members_count": "Members_count",
}


def bootstrap_sample_input(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return
    df = pd.DataFrame(
        [
            {
                "design_code": "AISC",
                "frame_type": "ClearSpan",
                "span_m": 30.0,
                "eave_height_m": 8.0,
                "bay_spacing_m": 6.0,
                "dead_load_kPa": 0.5,
                "live_load_kPa": 0.57,
                "collateral_kPa": 0.24,
                "wind_speed_mps": 45.0,
                "seismic_zone": "II",
                "steel_grade": "ASTM A572 Gr50",
                "Members_count": 5,
                # User-friendly columns present in the screenshot
                "Building type": "Multi span 1",
                "Width in m (o/o of steel)": 55.92,
                "Right eave height in m": 9.44,
                "Right end wall bay spacing in m": 7.31,
                "Side wall bay spacing": "1@7.985+5@7.99+1@7.985",
            }
        ]
    )
    with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name=INPUT_SHEET_NAME, index=False)


def _normalize_input_columns(df: pd.DataFrame) -> pd.DataFrame:
    # Map known synonyms to our internal names while preserving the originals for pass-through fields
    for col in list(df.columns):
        if col in SYNONYMS:
            target = SYNONYMS[col]
            if target not in df.columns:
                df[target] = df[col]
    return df


def read_input(path: str | Path) -> InputModel:
    p = Path(path)
    bootstrap_sample_input(p)
    xl = pd.ExcelFile(p)
    sheet = INPUT_SHEET_NAME if INPUT_SHEET_NAME in xl.sheet_names else xl.sheet_names[0]
    df = pd.read_excel(xl, sheet_name=sheet)

    df = _normalize_input_columns(df)

    # Validate minimum core fields needed for generation
    missing = [c for c in ["design_code", "frame_type", "span_m", "Members_count"] if c not in df.columns]
    if missing:
        raise ValueError(f"Input sheet missing required columns: {missing}")

    row = df.iloc[0].to_dict()
    return InputModel(**row)


MEMBER_TABLE_COLUMNS = [
    "Mark",
    "Web Start Depth (mm)",
    "Web End Depth (mm)",
    "Web Thickness (mm)",
    "Web Plate Length (mm)",
    "Outside Flange Width (mm)",
    "Outside Flange Thickness (mm)",
    "Outside Flange Length (mm)",
    "Inside Flange Width (mm)",
    "Inside Flange Thickness (mm)",
    "Inside Flange Length (mm)",
    "Weight (kg/m)",
    "Status",
]


def write_output(path: str | Path, members: List[MemberCandidate]) -> None:
    from openpyxl.styles import Alignment, Font

    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)

    data = [
        {
            "Member id": m.member_id,
            "Web start depth": m.web_start_depth_mm,
            "Web end depth": m.web_end_depth_mm,
            "Web thickness": m.web_thickness_mm,
            "Inside flange width": m.if_width_mm,
            "Inside flange thickness": m.if_thickness_mm,
            "Outside flange width": m.of_width_mm,
            "Outside flange thickness": m.of_thickness_mm,
            "Weight (kg/m)": round(m.weight_kg_per_m, 3),
            "Status": m.status,
        }
        for m in members
    ]
    df = pd.DataFrame(data)

    with pd.ExcelWriter(p, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=OUTPUT_SHEET_NAME, index=False)
        ws = writer.book[OUTPUT_SHEET_NAME]

        # Style header (first row)
        from openpyxl.styles import Alignment, Font
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center

        # Number formats: data rows start at row 2
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=8):
            for cell in row:
                cell.number_format = "0"
        for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
            for cell in row:
                cell.number_format = "0.00"


def write_member_table(path: str | Path, members: List[MemberCandidate], side_wall_bay_expr: Optional[str]) -> None:
    """Write a simplified 'Member Table' sheet.

    If a bay pattern is provided, expand rows per bay using that length (mm). Otherwise
    write one row per member with length = span_m (handled by caller if desired).
    """
    if not side_wall_bay_expr:
        return

    try:
        bays = parse_bay_pattern(side_wall_bay_expr)
    except Exception:
        return

    rows: List[dict] = []
    for m in members:
        for idx, bay_mm in enumerate(bays.bays_mm, start=1):
            rows.append(
                {
                    "Mark": f"{m.member_id}-{idx}",
                    "Web Start/End": f"{m.web_start_depth_mm} / {m.web_end_depth_mm}",
                    "Web Thick (mm)": m.web_thickness_mm,
                    "Web Length (mm)": bay_mm,
                    "Outside Flange (W x Thk x L)": f"{int(m.of_width_mm)} x {int(m.of_thickness_mm)} x {bay_mm}",
                    "Inside Flange (W x Thk x L)": f"{int(m.if_width_mm)} x {int(m.if_thickness_mm)} x {bay_mm}",
                }
            )

    df = pd.DataFrame(
        rows,
        columns=[
            "Mark",
            "Web Start/End",
            "Web Thick (mm)",
            "Web Length (mm)",
            "Outside Flange (W x Thk x L)",
            "Inside Flange (W x Thk x L)",
        ],
    )

    p = Path(path)
    with pd.ExcelWriter(p, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=MEMBER_TABLE_SHEET, index=False)


def write_member_table_final(path: str | Path, rows: List[dict]) -> None:
    from openpyxl.styles import Alignment, Font

    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows, columns=MEMBER_TABLE_COLUMNS)
    with pd.ExcelWriter(p, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=OUTPUT_SHEET_NAME, index=False)
        ws = writer.book[OUTPUT_SHEET_NAME]
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center
        # Formats: integers for sizes/lengths, weight 2 decimals
        # Mark is col 1; numeric cols 2..12; weight is col 12
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=11):
            for cell in row:
                cell.number_format = "0"
        for row in ws.iter_rows(min_row=2, min_col=12, max_col=12):
            for cell in row:
                cell.number_format = "0.00"


def write_member_table_compact(path: str | Path, rows: List[dict]) -> None:
    """Write a compact 'Member Table' sheet matching the screenshot layout.

    Columns:
      - Mark
      - Web Start/End
      - Web Thick (mm)
      - Web Length (mm)
      - Outside Flange (W x Thk x L)
      - Inside Flange (W x Thk x L)
    """
    compact_rows: List[dict] = []
    for r in rows:
        compact_rows.append(
            {
                "Mark": r["Mark"],
                "Web Depth Start/End": f"{r['Web Start Depth (mm)']} / {r['Web End Depth (mm)']}",
                "Web Plate Thick (mm)": r["Web Thickness (mm)"],
                "Web Plate Length (mm)": r["Web Plate Length (mm)"],
                "Outside Flange W x Thk x Length": f"{r['Outside Flange Width (mm)']} x {r['Outside Flange Thickness (mm)']} x {r['Outside Flange Length (mm)']}",
                "Inside Flange W x Thk x Length": f"{r['Inside Flange Width (mm)']} x {r['Inside Flange Thickness (mm)']} x {r['Inside Flange Length (mm)']}",
            }
        )

    df = pd.DataFrame(
        compact_rows,
        columns=[
            "Mark",
            "Web Depth Start/End",
            "Web Plate Thick (mm)",
            "Web Plate Length (mm)",
            "Outside Flange W x Thk x Length",
            "Inside Flange W x Thk x Length",
        ],
    )

    p = Path(path)
    with pd.ExcelWriter(p, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=MEMBER_TABLE_SHEET, index=False)
        # Make Member Table the active sheet by default when opening
        wb = writer.book
        if MEMBER_TABLE_SHEET in wb.sheetnames:
            wb.active = wb.sheetnames.index(MEMBER_TABLE_SHEET)


def write_output_template_from_rows(path: str | Path, rows_in: List[dict]) -> None:
    """Write grouped two-row header sheet like the provided template image.

    Columns under groups:
      Member id | Web (mm): Start depth, End depth, Thickness | Inside flange (mm): Width, Thickness | Outside flange (mm): Width, Thickness
    """
    from openpyxl.styles import Alignment, Font

    # Build flat rows from final dicts
    rows = []
    for idx, r in enumerate(rows_in, start=1):
        rows.append(
            {
                "Member id": idx,
                "Start depth": r.get("Web Start Depth (mm)"),
                "End depth": r.get("Web End Depth (mm)"),
                "Thickness": r.get("Web Thickness (mm)"),
                "IF Width": r.get("Inside Flange Width (mm)"),
                "IF Thickness": r.get("Inside Flange Thickness (mm)"),
                "OF Width": r.get("Outside Flange Width (mm)"),
                "OF Thickness": r.get("Outside Flange Thickness (mm)"),
            }
        )

    df = pd.DataFrame(
        rows,
        columns=[
            "Member id",
            "Start depth",
            "End depth",
            "Thickness",
            "IF Width",
            "IF Thickness",
            "OF Width",
            "OF Thickness",
        ],
    )

    p = Path(path)
    with pd.ExcelWriter(p, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=TEMPLATE_SHEET_NAME, index=False, startrow=1)
        ws = writer.book[TEMPLATE_SHEET_NAME]

        # Insert top grouped header row
        ws.cell(row=1, column=1, value="Member id")
        ws.cell(row=1, column=2, value="Web (mm)")
        ws.cell(row=1, column=5, value="Inside flange (mm)")
        ws.cell(row=1, column=7, value="Outside flange (mm)")

        # Merge group headers
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)
        ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=8)

        # Style
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center
        for cell in ws[2]:
            cell.font = header_font
            cell.alignment = center

        # Number formats for data rows (start at row 3)
        for row in ws.iter_rows(min_row=3, min_col=2, max_col=8):
            for cell in row:
                cell.number_format = "0"


def write_output_template(path: str | Path, members: List[MemberCandidate]) -> None:
    """Write a second sheet that matches the required grouped-header template.

    Top header row has merged labels: Web (mm), Inside flange (mm), Outside flange (mm).
    Second header row has sublabels. First column 'Member id' is merged vertically.
    """
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    p = Path(path)

    # Build flat data for the second row headers
    rows = []
    for idx, m in enumerate(members, start=1):
        rows.append(
            {
                "Member id": idx,
                "Start depth": m.web_start_depth_mm,
                "End depth": m.web_end_depth_mm,
                "Thickness": m.web_thickness_mm,
                "IF Width": m.if_width_mm,
                "IF Thickness": m.if_thickness_mm,
                "OF Width": m.of_width_mm,
                "OF Thickness": m.of_thickness_mm,
            }
        )
    df = pd.DataFrame(
        rows,
        columns=[
            "Member id",
            "Start depth",
            "End depth",
            "Thickness",
            "IF Width",
            "IF Thickness",
            "OF Width",
            "OF Thickness",
        ],
    )

    with pd.ExcelWriter(p, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=TEMPLATE_SHEET_NAME, index=False, startrow=1)
        ws = writer.book[TEMPLATE_SHEET_NAME]

        # Insert top grouped header row
        ws.cell(row=1, column=1, value="Member id")
        ws.cell(row=1, column=2, value="Web (mm)")
        ws.cell(row=1, column=5, value="Inside flange (mm)")
        ws.cell(row=1, column=7, value="Outside flange (mm)")

        # Merge group headers
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)
        ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=8)

        # Style
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center
        for cell in ws[2]:
            cell.font = header_font
            cell.alignment = center

        # Number formats for data rows (start at row 3)
        for row in ws.iter_rows(min_row=3, min_col=2, max_col=8):
            for cell in row:
                cell.number_format = "0"
