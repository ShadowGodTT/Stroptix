import openpyxl
import pandas as pd
import logging
import os
import sys
from pathlib import Path
from difflib import get_close_matches

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('excel_to_csv_debug.log'),
        logging.StreamHandler(sys.stdout)
    ]
)

def extract_parameter_value_pairs(sheet):
    """Extract parameter-value pairs from a sheet with optimized processing."""
    pairs = []
    logging.info(f"\nProcessing sheet: {sheet.title}")
    print(f"\n=== {sheet.title} ===")
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        if not any(cell.value for cell in row):
            continue
        for col_idx in range(1, len(row)-1, 2):
            if col_idx + 1 < len(row):
                param = str(row[col_idx].value).strip() if row[col_idx].value else ""
                value = str(row[col_idx+1].value).strip() if row[col_idx+1].value else ""
                if param.startswith('Sr') or not param or not value:
                    continue
                pairs.append((param, value))
                print(f"{param}: {value}")
    logging.info(f"Total pairs found in sheet {sheet.title}: {len(pairs)}")
    return pairs

def process_excel_file(input_file):
    """Process the Excel file and convert it to CSV and MBS formats (extract all parameters)."""
    try:
        logging.info(f"Attempting to read Excel file: {input_file}")
        print(f"\nReading Excel file: {input_file}")
        workbook = openpyxl.load_workbook(input_file, data_only=True, read_only=True)
        logging.info(f"Excel file loaded. Sheets found: {workbook.sheetnames}")
        all_pairs = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            pairs = extract_parameter_value_pairs(sheet)
            all_pairs.extend(pairs)
        df = pd.DataFrame(all_pairs, columns=['Parameter', 'Value'])
        total_params = len(all_pairs)
        logging.info(f"Total parameters found: {total_params}")
        print(f"\nTotal parameters found: {total_params}")
        exe_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        output_dir = os.path.join(exe_dir, 'output')
        os.makedirs(output_dir, exist_ok=True)
        input_filename = os.path.splitext(os.path.basename(input_file))[0]
        output_path = os.path.join(output_dir, f"{input_filename}_filtered.csv")
        df.to_csv(output_path, index=False, mode='w', encoding='utf-8')
        logging.info(f"Successfully saved CSV file: {output_path}")
        mbs_converter = MBSConverter(df)
        mbs_file = mbs_converter.generate_in_file()
        logging.info(f"Successfully generated MBS file: {mbs_file}")
        return output_path, mbs_file
    except Exception as e:
        logging.error(f"Error processing Excel file: {str(e)}")
        raise
    finally:
        if 'workbook' in locals():
            workbook.close()

class MBSConverter:
    def __init__(self, df):
        self.df = df
        self.output_dir = Path('output')
        self.output_dir.mkdir(exist_ok=True)
        self.sections = {}
        crane_capacity = self.get_value("Crane Capacity in MT")
        self.has_crane_data = crane_capacity is not None and float(crane_capacity) > 0
    def get_value(self, param_name, default=None):
        exact_match = self.df[self.df['Parameter'] == param_name]
        if not exact_match.empty:
            value = exact_match.iloc[0]['Value']
            if value == '-' or value == 'NA' or value == '':
                return default
            return value
        all_params = self.df['Parameter'].tolist()
        matches = get_close_matches(param_name, all_params, n=1, cutoff=0.8)
        if matches:
            value = self.df[self.df['Parameter'] == matches[0]].iloc[0]['Value']
            if value == '-' or value == 'NA' or value == '':
                return default
            return value
        return default
    def safe_float(self, value, default=0.0):
        if value == '-' or value == 'NA' or value == '':
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default
    # ... (other methods from MBSConverter should be copied as needed) ... 